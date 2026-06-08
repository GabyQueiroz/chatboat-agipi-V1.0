import csv
import os
from dotenv import load_dotenv
from pathlib import Path
import numpy as np
from openpyxl import load_workbook

from src.retrieval.embeddings import Embedder
from src.retrieval.vector_db import VectorStore
import src.core.rag_pipeline


load_dotenv()

ROOT_DIR = Path(__file__).resolve().parents[2]

EMBED_MODEL = os.getenv("EMBED_MODEL", "sentence-transformers/all-MiniLM-L6-v2")
relative_dataset_path = os.getenv("DATASET_PATH", "src/evaluation/PERGUNTAS_TESTE.xlsx")

DATASET_PATH = str(ROOT_DIR / relative_dataset_path)


def load_golden_dataset(xlsx_path: str):
    """Lê a planilha de testes e retorna um dicionário de Perguntas -> Documentos Esperados"""
    dataset = []

    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]

    if 'Pergunta' not in header or 'Base documental' not in header:
        wb.close()
        raise ValueError("A planilha precisa conter as colunas 'Pergunta' e 'Base documental'.")
        
    pergunta_idx = header.index('Pergunta')
    base_doc_idx = header.index('Base documental')

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(pergunta_idx, base_doc_idx):
            continue

        pergunta = row[pergunta_idx]
        base_doc_str = row[base_doc_idx]

        if pergunta is not None and base_doc_str is not None:
            pergunta_clean = str(pergunta).strip()
            base_doc_clean = str(base_doc_str).strip()

            if pergunta_clean and base_doc_clean:
                    expected_docs = [doc.strip() for doc in base_doc_clean.split(';')]
                    dataset.append({
                        "question": pergunta_clean,
                        "expected_docs": expected_docs
                    })

    wb.close()
    return dataset


def calculate_precision_at_k(retrieved_docs: list[str], expected_docs: list[str], k: int) -> float:
    """Calcula quantos dos top K documentos recuperados são relevantes."""
    top_k_retrieved = retrieved_docs[:k]
    
    # Conta quantos dos recuperados estão na lista de esperados
    hits = sum(1 for doc in top_k_retrieved if doc in expected_docs)
    
    return hits / k


def calculate_recall_at_k(retrieved_docs: list[str], expected_docs: list[str], k: int) -> float:
    """Calcula a proporção de documentos esperados que foram recuperados dentro dos top K."""
    top_k_retrieved = retrieved_docs[:k]
    if not expected_docs:
        return 0.0
    
    # Conta quantos dos documentos esperados foram encontrados no top K recuperado
    hits = sum(1 for doc in expected_docs if doc in top_k_retrieved)
    return hits / len(expected_docs)


def calculate_hit_at_k(retrieved_docs: list[str], expected_docs: list[str], k: int) -> float:
    """Verifica se pelo menos UM documento esperado está nos top K."""
    top_k_retrieved = retrieved_docs[:k]
    for doc in expected_docs:
        if doc in top_k_retrieved:
            return 1.0
    return 0.0


def calculate_mrr_at_k(retrieved_docs: list[str], expected_docs: list[str], k: int) -> float:
    """Calcula o Reciprocal Rank (1/rank) baseado na primeira ocorrência útil dentro do top K."""
    top_k_retrieved = retrieved_docs[:k]
    for rank, doc in enumerate(top_k_retrieved, start=1):
        if doc in expected_docs:
            return 1.0 / rank
    return 0.0


def run_evaluation():
    embedder = Embedder(model_name=EMBED_MODEL)
    vector_store = VectorStore(dimension=embedder.dimension)

    # Define e carrega o cache do banco vetorial estruturado idêntico ao main.py
    DATA_DIR = ROOT_DIR / "data"
    INDEX_CACHE_PATH = DATA_DIR / "index" / "faiss_index.bin"
    METADATA_CACHE_PATH = DATA_DIR / "processed" / "faiss_metadata.json"
    CACHE_THRESHOLD = 0.75

    vector_store.load(str(INDEX_CACHE_PATH), str(METADATA_CACHE_PATH))
    
    dataset = load_golden_dataset(DATASET_PATH)
    K = 50
    
    total_precision = 0.0
    total_recall = 0.0
    total_hit = 0.0
    total_mrr = 0.0
    total_cache_hits = 0.0
    
    print(f"Iniciando avaliação de {len(dataset)} perguntas...") 
    
    for item in dataset:
        question = item['question']
        print(f"QUESTION: {question}")
        expected_docs = item['expected_docs']
        print(f"EXPECTED DOCS: {expected_docs}")
        
        # Recupera os documentos do banco vetorial
        query_embedding = embedder.embed_texts([question])
        search_results = vector_store.search(np.array(query_embedding), top_k=K)


        is_cache_hit = 0.0
        if search_results:
            top_result = search_results[0]
            # Conta como cache hit se o Top 1 for FAQ e tiver score acima do threshold
            if top_result.get("doc_type") == "faq" and top_result.get("score", 0.0) >= CACHE_THRESHOLD:
                is_cache_hit = 1.0
        
        total_cache_hits += is_cache_hit

        filtered_results = [res for res in search_results if res.get("doc_type") != "faq"]
        # print(f"FILTERED SEARCH RESULTS: {filtered_results[:1]}")
        
        # Extrai os nomes dos documentos recuperados
        retrieved_docs = [os.path.basename(res.get("source", "")) for res in filtered_results[:K]]
        print(f"RETRIEVED DOCS: {retrieved_docs}")
        
        # Calcula as métricas para esta pergunta
        precision = calculate_precision_at_k(retrieved_docs, expected_docs, K)
        recall = calculate_recall_at_k(retrieved_docs, expected_docs, K)
        hit = calculate_hit_at_k(retrieved_docs, expected_docs, K)
        mrr = calculate_mrr_at_k(retrieved_docs, expected_docs, K)
        
        total_precision += precision
        total_recall += recall
        total_hit += hit
        total_mrr += mrr
        
        print(f"Q: {question[:40]}... | P@{K}: {precision:.2f} | R@{K}: {recall:.2f} | Hit@{K}: {hit:.2f} | MRR@{K}: {mrr:.2f}")

    # Média final do sistema
    num_queries = len(dataset)
    mean_precision = total_precision / num_queries if num_queries > 0 else 0
    mean_recall = total_recall / num_queries if num_queries > 0 else 0
    mean_hit = total_hit / num_queries if num_queries > 0 else 0
    mean_mrr = total_mrr / num_queries if num_queries > 0 else 0
    mean_cache_hit_rate = total_cache_hits / num_queries if num_queries > 0 else 0
    
    print(f"\n[RESULTADO FINAL]")
    print(f"-> Cache Hit Rate (FAQ): {mean_cache_hit_rate:.2%}")
    print(f"-> Precision@{K} Médio do Sistema: {mean_precision:.2f}")
    print(f"-> Recall@{K} Médio do Sistema: {mean_recall:.2f}")
    print(f"-> Hit@{K} (Hit Rate) Médio do Sistema: {mean_hit:.2f}")
    print(f"-> MRR@{K} Médio: {mean_mrr:.2f}")

if __name__ == "__main__":
    run_evaluation()