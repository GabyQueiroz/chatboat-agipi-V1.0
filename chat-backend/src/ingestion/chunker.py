import os
import pickle
import re
from pathlib import Path
from typing import Any

import pymupdf4llm
from docx import Document
from openpyxl import load_workbook


def chunk_text(text: str, chunk_size: int = 900, overlap: int = 120) -> list[str]:
    chunks = []
    start = 0
    text = normalize_text(text)
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += max(chunk_size - overlap, 1)
    return chunks


def normalize_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"==>\s*picture.*?<==", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"-{3,}\s*Start of picture text\s*-{3,}", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"-{3,}\s*End of picture text\s*-{3,}", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b[pP]icture\b", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def save_documents(documents: list[dict[str, Any]], cache_path: str) -> None:
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    with open(cache_path, "wb") as file:
        pickle.dump(documents, file)
    print(f"[CHUNKER] Documentos salvos em cache: {cache_path}")


def load_documents(cache_path: str) -> list[dict[str, Any]] | None:
    if not os.path.exists(cache_path):
        return None
    try:
        with open(cache_path, "rb") as file:
            documents = pickle.load(file)
        print(f"[CHUNKER] Documentos carregados do cache: {cache_path}")
        return documents
    except Exception as exc:
        print(f"[CHUNKER] Erro ao carregar cache: {exc}")
        return None


def save_source_manifest(manifest: list[dict[str, Any]], manifest_path: str) -> None:
    os.makedirs(os.path.dirname(manifest_path), exist_ok=True)
    with open(manifest_path, "wb") as file:
        pickle.dump(manifest, file)
    print(f"[CHUNKER] Manifest salvo em: {manifest_path}")


def load_source_manifest(manifest_path: str) -> list[dict[str, Any]] | None:
    if not os.path.exists(manifest_path):
        return None
    try:
        with open(manifest_path, "rb") as file:
            manifest = pickle.load(file)
        print(f"[CHUNKER] Manifest carregado de: {manifest_path}")
        return manifest
    except Exception as exc:
        print(f"[CHUNKER] Erro ao carregar manifest: {exc}")
        return None


def build_source_manifest(raw_dirs: list[str], faq_path: str | None = None) -> list[dict[str, Any]]:
    manifest: list[dict[str, Any]] = []

    for raw_dir in raw_dirs:
        root = Path(raw_dir)
        if not root.exists():
            continue
        for path in sorted(root.rglob("*")):
            if not path.is_file():
                continue
            if path.suffix.lower() not in {".pdf", ".txt", ".docx"}:
                continue
            stat = path.stat()
            manifest.append(
                {
                    "kind": "document",
                    "path": str(path.resolve()),
                    "size": stat.st_size,
                    "mtime": int(stat.st_mtime),
                }
            )

    if faq_path:
        faq_file = Path(faq_path)
        if faq_file.exists():
            stat = faq_file.stat()
            manifest.append(
                {
                    "kind": "faq",
                    "path": str(faq_file.resolve()),
                    "size": stat.st_size,
                    "mtime": int(stat.st_mtime),
                }
            )

    return manifest


def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def read_docx_file(path: Path) -> str:
    document = Document(str(path))
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(paragraphs)


def read_pdf_file(path: Path) -> str:
    return pymupdf4llm.to_text(str(path))


def read_document_content(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf_file(path)
    if suffix == ".docx":
        return read_docx_file(path)
    return read_text_file(path)


def process_document_directories(raw_dirs: list[str]) -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    allowed_suffixes = {".pdf", ".txt", ".docx"}

    for raw_dir in raw_dirs:
        root = Path(raw_dir)
        if not root.exists():
            print(f"[CHUNKER] Pasta nao encontrada: {root}")
            continue

        for file_path in sorted(root.rglob("*")):
            if not file_path.is_file() or file_path.suffix.lower() not in allowed_suffixes:
                continue

            try:
                relative_path = file_path.relative_to(root).as_posix()
                source_label = f"{root.name}/{relative_path}"
                print(f"[CHUNKER] Processando arquivo: {source_label}")
                content = normalize_text(read_document_content(file_path))
                chunks = chunk_text(content)

                for index, chunk in enumerate(chunks):
                    documents.append(
                        {
                            "id": f"{file_path.stem}_chunk_{index}",
                            "source": source_label,
                            "title": file_path.stem,
                            "text": chunk,
                            "chunk_index": index,
                            "doc_type": "document",
                            "category": file_path.parent.name,
                        }
                    )

                print(f"[CHUNKER] Extraidos {len(chunks)} chunks de {source_label}")
            except Exception as exc:
                print(f"[CHUNKER] Falha ao processar {file_path}: {exc}")

    return documents


def process_faq_workbook(faq_path: str | None) -> list[dict[str, Any]]:
    if not faq_path:
        return []

    workbook_path = Path(faq_path)
    if not workbook_path.exists():
        print(f"[CHUNKER] FAQ nao encontrado: {workbook_path}")
        return []

    workbook = load_workbook(str(workbook_path), read_only=True, data_only=True)
    documents: list[dict[str, Any]] = []
    doc_index = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_map = None

        for row in rows:
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            lower_values = [value.lower() for value in values]

            if header_map is None and "pergunta" in lower_values and "resposta" in lower_values:
                header_map = {name.lower(): index for index, name in enumerate(lower_values)}
                continue

            if header_map is None:
                continue

            question = values[header_map.get("pergunta", -1)] if header_map.get("pergunta", -1) >= 0 else ""
            answer = values[header_map.get("resposta", -1)] if header_map.get("resposta", -1) >= 0 else ""
            category = values[header_map.get("categoria", -1)] if header_map.get("categoria", -1) >= 0 else sheet_name
            evidence = values[header_map.get("evidência", -1)] if header_map.get("evidência", -1) >= 0 else ""
            base_documental = values[header_map.get("base documental", -1)] if header_map.get("base documental", -1) >= 0 else ""

            if not question or not answer:
                continue

            text = normalize_text(
                f"Pergunta: {question}\nCategoria: {category}\nResposta: {answer}\n"
                f"Evidencia: {evidence}\nBase documental: {base_documental}"
            )
            documents.append(
                {
                    "id": f"faq_{doc_index}",
                    "source": f"FAQ/{sheet_name}",
                    "title": question,
                    "text": text,
                    "chunk_index": 0,
                    "doc_type": "faq",
                    "category": category or sheet_name,
                    "faq_question": question,
                    "faq_answer": answer,
                    "evidence": evidence,
                    "base_documental": base_documental,
                }
            )
            doc_index += 1

    print(f"[CHUNKER] FAQ processado com {len(documents)} entradas")
    return documents


def process_documents(raw_dirs: list[str], faq_path: str | None = None) -> list[dict[str, Any]]:
    documents = []
    documents.extend(process_faq_workbook(faq_path))
    documents.extend(process_document_directories(raw_dirs))
    return documents
